"""Export functionality for DirText - TXT, CSV, JSON, XLSX export"""

import csv
import ctypes
import fnmatch
import json
import os
from datetime import datetime

from PyQt6.QtWidgets import (
    QDialog, QLabel, QVBoxLayout, QHBoxLayout,
    QCheckBox, QFileDialog, QMessageBox, QWidget,
)
from animations import AnimatedButton


# ---------- 格式选项（多选复选框） / Format Check Option ----------
class FormatCheckOption(QWidget):
    """可点击的复选框选项，显示 ☐ / ☑"""

    def __init__(self, text, C, parent=None):
        super().__init__(parent)
        self._C = C
        self._checked = False
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 2, 0, 2)
        layout.setSpacing(6)

        self._box = QLabel()
        self._box.setFixedWidth(16)
        self._text = QLabel(text)
        self._text.setStyleSheet(f"color: {C['fg']}; font-size: 10pt; background: transparent;")
        layout.addWidget(self._box)
        layout.addWidget(self._text)
        layout.addStretch()
        self._update_display()

    def _update_display(self):
        C = self._C
        if self._checked:
            self._box.setText(f'<span style="color: {C["accent"]}; font-size: 13pt;">☑</span>')
        else:
            self._box.setText(f'<span style="color: {C["option_circle"]}; font-size: 13pt;">☐</span>')

    def set_checked(self, v):
        self._checked = v
        self._update_display()

    def is_checked(self):
        return self._checked

    def mousePressEvent(self, event):
        self._checked = not self._checked
        self._update_display()


# ---------- 导出格式对话框（多选） / Export Format Dialog ----------
class ExportFormatDialog(QDialog):
    """多选导出格式对话框，可同时勾选多种格式批量导出。"""

    def __init__(self, parent=None, t=None, C=None, make_btn_style=None):
        super().__init__(parent)
        self._t = t or (lambda k, **kw: k)
        self._C = C or {}
        self._make_btn_style = make_btn_style or (lambda c: '')
        self.selected_formats = {'txt'}  # default selection
        self._setup_ui()

    def _setup_ui(self):
        t = self._t
        C = self._C
        self.setWindowTitle(t('export_format'))
        self.setFixedSize(300, 250)
        self.setStyleSheet(f"QDialog {{ background: {C['bg']}}};")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 15, 20, 10)
        layout.setSpacing(8)

        prompt = QLabel(t('format_prompt'))
        prompt.setStyleSheet(f"color: {C['fg']}; font-size: 10pt; background: transparent;")
        layout.addWidget(prompt)

        self.opt_txt = FormatCheckOption(t('format_txt'), C)
        self.opt_txt.set_checked(True)
        layout.addWidget(self.opt_txt)

        self.opt_csv = FormatCheckOption(t('format_csv'), C)
        layout.addWidget(self.opt_csv)

        self.opt_json = FormatCheckOption(t('format_json'), C)
        layout.addWidget(self.opt_json)

        self.opt_xlsx = FormatCheckOption(t('format_xlsx'), C)
        layout.addWidget(self.opt_xlsx)

        layout.addSpacing(6)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        for text, slot in [('OK', self._accept), ('Cancel', self.reject)]:
            btn = AnimatedButton(text)
            btn.setStyleSheet(self._make_btn_style(C))
            btn.clicked.connect(slot)
            btn_layout.addWidget(btn)
        layout.addLayout(btn_layout)

    def _accept(self):
        checked = []
        if self.opt_txt.is_checked():
            checked.append('txt')
        if self.opt_csv.is_checked():
            checked.append('csv')
        if self.opt_json.is_checked():
            checked.append('json')
        if self.opt_xlsx.is_checked():
            checked.append('xlsx')

        if not checked:
            t = self._t
            QMessageBox.warning(self, 'Warning', t('no_format_selected'))
            return

        self.selected_formats = checked
        self.accept()


# ---------- 元数据对话框 / Metadata Dialog ----------
class MetadataDialog(QDialog):
    def __init__(self, parent=None, current_options=None, t=None, C=None, make_btn_style=None):
        super().__init__(parent)
        self._t = t or (lambda k, **kw: k)
        self._C = C or {}
        self._make_btn_style = make_btn_style or (lambda c: '')
        self.options = dict(current_options) if current_options else {
            'size': False, 'created': False, 'modified': False, 'accessed': False,
        }
        self._setup_ui()

    def _setup_ui(self):
        t = self._t
        C = self._C
        self.setWindowTitle(t('metadata_title'))
        self.setFixedSize(320, 280)
        self.setStyleSheet(f"QDialog {{ background: {C['bg']}}};")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 15, 20, 10)
        layout.setSpacing(8)

        prompt = QLabel(t('metadata_prompt'))
        prompt.setStyleSheet(f"color: {C['fg']}; font-size: 10pt; background: transparent;")
        layout.addWidget(prompt)

        hint = QLabel(t('metadata_txt_hint'))
        hint.setStyleSheet(f"color: {C['fg2']}; font-size: 9pt; background: transparent;")
        hint.setWordWrap(True)
        layout.addWidget(hint)

        checkbox_style = f"color: {C['fg']}; font-size: 10pt; background: transparent;"

        self.cb_size = QCheckBox(t('metadata_size'))
        self.cb_size.setChecked(self.options.get('size', False))
        self.cb_size.setStyleSheet(checkbox_style)
        layout.addWidget(self.cb_size)

        self.cb_created = QCheckBox(t('metadata_created'))
        self.cb_created.setChecked(self.options.get('created', False))
        self.cb_created.setStyleSheet(checkbox_style)
        layout.addWidget(self.cb_created)

        self.cb_modified = QCheckBox(t('metadata_modified'))
        self.cb_modified.setChecked(self.options.get('modified', False))
        self.cb_modified.setStyleSheet(checkbox_style)
        layout.addWidget(self.cb_modified)

        self.cb_accessed = QCheckBox(t('metadata_accessed'))
        self.cb_accessed.setChecked(self.options.get('accessed', False))
        self.cb_accessed.setStyleSheet(checkbox_style)
        layout.addWidget(self.cb_accessed)

        layout.addSpacing(6)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        for text, slot in [('OK', self._accept), ('Cancel', self.reject)]:
            btn = AnimatedButton(text)
            btn.setStyleSheet(self._make_btn_style(C))
            btn.clicked.connect(slot)
            btn_layout.addWidget(btn)
        layout.addLayout(btn_layout)

    def _accept(self):
        self.options = {
            'size': self.cb_size.isChecked(),
            'created': self.cb_created.isChecked(),
            'modified': self.cb_modified.isChecked(),
            'accessed': self.cb_accessed.isChecked(),
        }
        self.accept()


# ── 小型工具 ──────────────────────────────────────────────
def _throttle_progress_step(total):
    """Return a step interval that yields ~200 progress updates for large datasets."""
    if total <= 200:
        return 1
    return total // 200


# ---------- 导出器 / Exporter ----------
class Exporter:
    """导出器，封装 TXT / CSV / JSON / XLSX 导出的完整流程。

    通过依赖注入接收工具函数和 app 引用，避免对模块级全局变量的直接耦合。
    """

    def __init__(self, app, t, lang, format_size, format_timestamp,
                 get_filtered_entries, get_structured_data,
                 get_full_preview_text, get_scan_total_dirs,
                 get_scan_total_files):
        self.app = app
        self.t = t
        self.lang = lang
        self.format_size = format_size
        self.format_timestamp = format_timestamp
        self._get_filtered_entries = get_filtered_entries
        self._get_structured_data = get_structured_data
        self._get_full_preview_text = get_full_preview_text
        self._get_scan_total_dirs = get_scan_total_dirs
        self._get_scan_total_files = get_scan_total_files

    # ── 内部辅助 ───────────────────────────────────────────

    def _progress_throttle(self, total):
        """Return (step, final_value) for progress bar updates."""
        if total <= 200:
            return 1, total
        return total // 200, total

    # ── TXT export ─────────────────────────────────────────

    def export_txt(self, timestamp):
        t = self.t
        fname, _ = QFileDialog.getSaveFileName(
            self.app, t('save_title'),
            t('file_name', ts=timestamp),
            'Text files (*.txt);;All files (*.*)',
        )
        if not fname:
            return
        try:
            el = self.app.export_lang
            entries = self._get_filtered_entries()
            if entries != self._get_structured_data():
                text = self._build_txt_from_entries(entries, el)
            elif el != self.lang:
                text = self._build_txt_content(el)
            else:
                text = self._get_full_preview_text()
            with open(fname, 'w', encoding='utf-8') as f:
                f.write(text + f"\n\nGenerated: {datetime.now():%Y-%m-%d %H:%M:%S}")
            QMessageBox.information(self.app, 'Success', t('success', path=fname))
            self.app.status.showMessage(os.path.basename(fname))
            self.open_export_folder(fname)
        except Exception as e:
            QMessageBox.critical(self.app, 'Error', t('fail', error=str(e)))

    def _build_txt_content(self, lang):
        sep = '═' * 80
        folders = self.app.folders
        total_dirs = self._get_scan_total_dirs()
        total_files = self._get_scan_total_files()
        lines = []
        for i, folder in enumerate(folders):
            lines.append(sep)
            lines.append(f"Folder [{i + 1}]: {folder}" if lang == 'en'
                         else f"文件夹 [{i + 1}]：{folder}")
            lines.append(sep)
            self._build_tree_lines(folder, self.app.recursion_depth, 0, '', lang, lines)
            lines.append('')
        lines.append(sep)
        if lang == 'en':
            lines.append(f"Total: {len(folders)} folder(s)")
            lines.append(f"         {total_dirs} folders, {total_files} files")
        else:
            lines.append(f"总计：{len(folders)} 个文件夹")
            lines.append(f"         共 {total_dirs} 个文件夹，{total_files} 个文件")
        return '\n'.join(lines)

    def _build_tree_lines(self, folder, max_depth, current_depth, prefix, lang, lines):
        try:
            entries = sorted(os.scandir(folder), key=lambda e: e.name)
            ignore = self.app.ignore_patterns
            if ignore:
                entries = [e for e in entries
                           if not any(fnmatch.fnmatch(e.name, p) for p in ignore)]
        except PermissionError:
            msg = '<Access denied>' if lang == 'en' else '<无法访问>'
            lines.append(f"{prefix}└── {msg}")
            return
        except Exception as e:
            lines.append(f"{prefix}└── <Error: {e}>")
            return

        for i, entry in enumerate(entries):
            is_last = (i == len(entries) - 1)
            connector = "└── " if is_last else "├── "
            child_prefix = "    " if is_last else "│   "
            if entry.is_dir():
                lines.append(f"{prefix}{connector}{entry.name}/")
                if max_depth == -1 or current_depth < max_depth:
                    self._build_tree_lines(entry.path, max_depth, current_depth + 1,
                                           prefix + child_prefix, lang, lines)
            else:
                lines.append(f"{prefix}{connector}{entry.name}")

    def _build_txt_from_entries(self, entries, lang):
        sep = '═' * 80
        folders = self.app.folders
        grouped = {}
        for e in entries:
            grouped.setdefault(e['folder'], []).append(e)

        lines = []
        for i, folder in enumerate(folders):
            items = grouped.get(folder)
            if items is None:
                continue
            lines.append(sep)
            lines.append(f"Folder [{i + 1}]: {folder}" if lang == 'en'
                         else f"文件夹 [{i + 1}]：{folder}")
            lines.append(sep)
            for entry in items:
                try:
                    rel = os.path.relpath(entry['path'], folder)
                except ValueError:
                    rel = entry['path']
                lines.append(f"  {rel}/" if entry['type'] == 'dir' else f"  {rel}")
            lines.append('')

        lines.append(sep)
        total_dirs = sum(1 for e in entries if e['type'] == 'dir')
        total_files = sum(1 for e in entries if e['type'] == 'file')
        if lang == 'en':
            lines.append(f"Total: {len(folders)} folder(s)")
            lines.append(f"         {total_dirs} folders, {total_files} files")
        else:
            lines.append(f"总计：{len(folders)} 个文件夹")
            lines.append(f"         共 {total_dirs} 个文件夹，{total_files} 个文件")
        return '\n'.join(lines)

    @staticmethod
    def open_export_folder(fname):
        folder = os.path.dirname(fname)
        desktop = os.path.normpath(os.path.join(os.path.expanduser('~'), 'Desktop'))
        if os.path.normpath(folder) == desktop:
            SHCNE_UPDATEDIR = 0x00001000
            SHCNF_PATHW = 0x0005
            ctypes.windll.shell32.SHChangeNotify(SHCNE_UPDATEDIR, SHCNF_PATHW, folder, None)
        else:
            os.startfile(folder)

    # ── CSV export ─────────────────────────────────────────

    def export_csv(self, timestamp):
        t = self.t
        fname, _ = QFileDialog.getSaveFileName(
            self.app, t('csv_save_title'),
            t('csv_file_name', ts=timestamp),
            'CSV files (*.csv);;All files (*.*)',
        )
        if not fname:
            return
        export_data = self._get_filtered_entries()
        total = len(export_data)

        pb = self.app.progress_bar
        pb.setRange(0, total)
        pb.setValue(0)
        pb.setFormat(t('export_progress'))
        pb.show()
        self.app.status.showMessage(t('exporting'))

        # Cache local refs
        el = self.app.export_lang
        meta = self.app.metadata_options
        _fmt_ts = self.format_timestamp
        has_size = meta.get('size')
        has_created = meta.get('created')
        has_modified = meta.get('modified')
        has_accessed = meta.get('accessed')

        # Pre-cache translated headers & labels
        col_folder = t('csv_col_folder', lang=el)
        col_name = t('csv_col_name', lang=el)
        col_path = t('csv_col_path', lang=el)
        col_type = t('csv_col_type', lang=el)
        col_level = t('csv_col_level', lang=el)
        type_dir = t('csv_type_dir', lang=el)
        type_file = t('csv_type_file', lang=el)

        headers = [col_folder, col_name, col_path, col_type, col_level]
        if has_size:
            headers.append(t('csv_col_size', lang=el))
        if has_created:
            headers.append(t('csv_col_created', lang=el))
        if has_modified:
            headers.append(t('csv_col_modified', lang=el))
        if has_accessed:
            headers.append(t('csv_col_accessed', lang=el))

        step = _throttle_progress_step(total)
        try:
            with open(fname, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(headers)

                for i, entry in enumerate(export_data):
                    row = [
                        entry['folder'], entry['name'], entry['path'],
                        type_dir if entry['type'] == 'dir' else type_file,
                        entry['level'],
                    ]
                    if has_size:
                        row.append(f"{entry['size']} B" if entry['type'] == 'file' else '')
                    if has_created:
                        row.append(_fmt_ts(entry['created']))
                    if has_modified:
                        row.append(_fmt_ts(entry['modified']))
                    if has_accessed:
                        row.append(_fmt_ts(entry['accessed']))
                    writer.writerow(row)
                    if i % step == 0:
                        pb.setValue(i + 1)
                pb.setValue(total)

            success_msg = t('success', path=fname)
            if has_created or has_modified or has_accessed:
                success_msg += '\n\n' + t('csv_excel_hint')
            QMessageBox.information(self.app, 'Success', success_msg)
            self.app.status.showMessage(os.path.basename(fname))
            self.open_export_folder(fname)
        except Exception as e:
            QMessageBox.critical(self.app, 'Error', t('fail', error=str(e)))
        finally:
            pb.hide()

    # ── JSON export ────────────────────────────────────────

    def export_json(self, timestamp):
        t = self.t
        fname, _ = QFileDialog.getSaveFileName(
            self.app, t('json_save_title'),
            t('json_file_name', ts=timestamp),
            'JSON files (*.json);;All files (*.*)',
        )
        if not fname:
            return
        export_data = self._get_filtered_entries()
        total = len(export_data)

        pb = self.app.progress_bar
        pb.setRange(0, total)
        pb.setValue(0)
        pb.setFormat(t('export_progress'))
        pb.show()
        self.app.status.showMessage(t('exporting'))

        # Cache local refs
        folders = self.app.folders
        el = self.app.export_lang
        meta = self.app.metadata_options
        _fmt_size = self.format_size
        _fmt_ts = self.format_timestamp
        has_size = meta.get('size')
        has_created = meta.get('created')
        has_modified = meta.get('modified')
        has_accessed = meta.get('accessed')

        # Pre-cache translated keys (avoid t() call per row)
        k_folder = t('csv_col_folder', lang=el)
        k_name = t('csv_col_name', lang=el)
        k_path = t('csv_col_path', lang=el)
        k_type = t('csv_col_type', lang=el)
        k_level = t('csv_col_level', lang=el)
        type_dir = t('csv_type_dir', lang=el)
        type_file = t('csv_type_file', lang=el)

        step = _throttle_progress_step(total)
        try:
            data = []
            for i, entry in enumerate(export_data):
                item = {
                    k_folder: entry['folder'],
                    k_name: entry['name'],
                    k_path: entry['path'],
                    k_type: type_dir if entry['type'] == 'dir' else type_file,
                    k_level: entry['level'],
                }
                if has_size:
                    item['size_bytes'] = entry['size'] if entry['type'] == 'file' else 0
                    item['size'] = _fmt_size(entry['size']) if entry['type'] == 'file' else '-'
                if has_created:
                    item['created'] = _fmt_ts(entry['created'])
                    item['created_timestamp'] = entry['created']
                if has_modified:
                    item['modified'] = _fmt_ts(entry['modified'])
                    item['modified_timestamp'] = entry['modified']
                if has_accessed:
                    item['accessed'] = _fmt_ts(entry['accessed'])
                    item['accessed_timestamp'] = entry['accessed']
                data.append(item)
                if i % step == 0:
                    pb.setValue(i + 1)
            pb.setValue(total)

            total_dirs = sum(1 for e in export_data if e['type'] == 'dir')
            total_files_count = sum(1 for e in export_data if e['type'] == 'file')

            export = {
                'export_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'folders': folders,
                'total_folders': len(folders),
                'total_dirs': total_dirs,
                'total_files': total_files_count,
                'entries': data,
            }
            with open(fname, 'w', encoding='utf-8') as f:
                json.dump(export, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self.app, 'Success', t('success', path=fname))
            self.app.status.showMessage(os.path.basename(fname))
            self.open_export_folder(fname)
        except Exception as e:
            QMessageBox.critical(self.app, 'Error', t('fail', error=str(e)))
        finally:
            pb.hide()

    # ── XLSX export (optimized) ────────────────────────────

    def export_xlsx(self, timestamp):
        t = self.t
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self.app, 'Error', t('xlsx_no_module'))
            return

        fname, _ = QFileDialog.getSaveFileName(
            self.app, t('xlsx_save_title'),
            t('xlsx_file_name', ts=timestamp),
            'Excel files (*.xlsx);;All files (*.*)',
        )
        if not fname:
            return

        export_data = self._get_filtered_entries()
        total = len(export_data)

        pb = self.app.progress_bar
        pb.setRange(0, total)
        pb.setValue(0)
        pb.setFormat(t('export_progress'))
        pb.show()
        self.app.status.showMessage(t('exporting'))

        # Cache local refs
        el = self.app.export_lang
        meta = self.app.metadata_options
        has_size = meta.get('size')
        has_created = meta.get('created')
        has_modified = meta.get('modified')
        has_accessed = meta.get('accessed')
        type_dir = t('csv_type_dir', lang=el)
        type_file = t('csv_type_file', lang=el)
        step = _throttle_progress_step(total)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'File List'

            # ── Pre-compute header row ──
            headers = [
                t('csv_col_folder', lang=el), t('csv_col_name', lang=el),
                t('csv_col_path', lang=el), t('csv_col_type', lang=el),
                t('csv_col_level', lang=el),
            ]
            if has_size:
                headers.append(t('csv_col_size', lang=el))
            if has_created:
                headers.append(t('csv_col_created', lang=el))
            if has_modified:
                headers.append(t('csv_col_modified', lang=el))
            if has_accessed:
                headers.append(t('csv_col_accessed', lang=el))

            ncols = len(headers)

            # ── Determine which data-column indices are dates (1-based) ──
            date_cols = set()
            col_idx = 6  # after the 5 fixed columns
            if has_size:
                col_idx += 1
            if has_created:
                date_cols.add(col_idx)
                col_idx += 1
            if has_modified:
                date_cols.add(col_idx)
                col_idx += 1
            if has_accessed:
                date_cols.add(col_idx)
                col_idx += 1

            # ── Style objects (created once, reused) ──
            header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=10)
            header_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # ── Write header row with styling ──
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # ── Write data rows via append() (fast path, no per-cell styling) ──
            for i, entry in enumerate(export_data):
                row_data = [
                    entry['folder'], entry['name'], entry['path'],
                    type_dir if entry['type'] == 'dir' else type_file,
                    entry['level'],
                ]
                if has_size:
                    row_data.append(f"{entry['size']} B" if entry['type'] == 'file' else '')
                # For date columns, pass formatted string so no cell format is needed
                if has_created:
                    ts = entry['created']
                    row_data.append(datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
                                    if ts else '')
                if has_modified:
                    ts = entry['modified']
                    row_data.append(datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
                                    if ts else '')
                if has_accessed:
                    ts = entry['accessed']
                    row_data.append(datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
                                    if ts else '')
                ws.append(row_data)
                if i % step == 0:
                    pb.setValue(i + 1)

            pb.setValue(total)

            # ── Post-processing: column widths (approximate, only header matters) ──
            for ci in range(1, ncols + 1):
                col_letter = get_column_letter(ci)
                header_len = len(str(ws.cell(row=1, column=ci).value or ''))
                ws.column_dimensions[col_letter].width = min(max(header_len + 3, 10), 65)

            # Freeze top row + auto-filter
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

            wb.save(fname)
            QMessageBox.information(self.app, 'Success', t('success', path=fname))
            self.app.status.showMessage(os.path.basename(fname))
            self.open_export_folder(fname)
        except Exception as e:
            QMessageBox.critical(self.app, 'Error', t('fail', error=str(e)))
        finally:
            pb.hide()
